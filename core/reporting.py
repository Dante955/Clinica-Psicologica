from sqlalchemy import func
from database.database_manager import session_scope
from database.models import Appointment, Expense, User, UserRole, Income
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
import os

class ReportingService:

    def get_financial_summary(self, start_date: datetime, end_date: datetime):
        """
        Calcula los ingresos, gastos y beneficios en un rango de fechas.
        """
        with session_scope() as session:
            # Calcular ingresos por citas (completadas y pagadas)
            income_from_appointments = session.query(func.sum(Appointment.price)).filter(
                Appointment.appointment_time.between(start_date, end_date),
                Appointment.status == 'completed',
                Appointment.payment_status == 'pagado'
            ).scalar()
            income_from_appointments = float(income_from_appointments) if income_from_appointments is not None else 0.0

            # Calcular otros ingresos desde la nueva tabla Income
            other_income = session.query(func.sum(Income.amount)).filter(
                Income.income_date.between(start_date, end_date)
            ).scalar()
            other_income = float(other_income) if other_income is not None else 0.0

            # Calcular gastos totales
            total_expenses = session.query(func.sum(Expense.amount)).filter(
                Expense.expense_date.between(start_date, end_date)
            ).scalar()
            total_expenses = float(total_expenses) if total_expenses is not None else 0.0

            total_income = income_from_appointments + other_income
            profit = total_income - total_expenses

            return {
                "income": total_income,
                "expenses": total_expenses,
                "profit": profit
            }

    def get_appointments_per_psychologist(self, start_date: datetime, end_date: datetime):
        """
        Devuelve un diccionario con el número de citas por psicólogo.
        """
        with session_scope() as session:
            # Contar citas por psicólogo
            results = session.query(
                User.username, 
                func.count(Appointment.id)
            ).join(Appointment, User.id == Appointment.psychologist_id).filter(
                User.role == UserRole.PSICOLOGO.value,  # Comparar con el valor string
                Appointment.appointment_time.between(start_date, end_date)
            ).group_by(User.username).all()

            # Convertir el resultado en un diccionario {psicologo: cantidad}
            return {username: count for username, count in results}

    def get_expenses_by_category(self, start_date: datetime, end_date: datetime, category: str = None):
        """
        Obtiene los gastos filtrados por categoría (test, terapia, medicación, etc.).
        Si category es None, devuelve todos los gastos.
        """
        with session_scope() as session:
            query = session.query(Expense).filter(
                Expense.expense_date.between(start_date, end_date)
            )
            
            if category:
                query = query.filter(Expense.expense_type == category)
            
            expenses = query.order_by(Expense.expense_date).all()
            return [(e.id, e.description, float(e.amount), e.expense_date, e.expense_type) for e in expenses]

    def export_financial_report_to_excel(self, start_date: datetime, end_date: datetime, output_path: str = None, selected_categories: list = None):
        """
        Exporta un reporte financiero detallado a Excel con información de gastos por categoría.
        Si selected_categories es None, incluye todas las categorías.
        """
        if output_path is None:
            output_path = f"reportes_financieros_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        
        # Crear un nuevo libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Gastos Financieros"
        
        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        category_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        category_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título del reporte
        ws['A1'] = "REPORTE FINANCIERO DE GASTOS"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        
        ws['A2'] = f"Del {start_date.strftime('%d/%m/%Y')} al {end_date.strftime('%d/%m/%Y')}"
        ws['A2'].font = Font(size=11, italic=True)
        ws.merge_cells('A2:E2')
        
        # Encabezados de tabla
        headers = ["Fecha", "Descripción", "Categoría", "Monto", "Notas"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Si no se especifican categorías, usar todas
        if selected_categories is None:
            selected_categories = ["general", "salario", "luz", "agua", "internet", "alquiler"]
        
        current_row = 5
        category_totals = {}
        
        with session_scope() as session:
            for category in selected_categories:
                expenses = session.query(Expense).filter(
                    Expense.expense_date.between(start_date, end_date),
                    Expense.expense_type == category
                ).order_by(Expense.expense_date).all()
                
                if expenses:
                    # Fila de categoría
                    ws.cell(row=current_row, column=1).value = category.upper()
                    for col in range(1, 6):
                        ws.cell(row=current_row, column=col).fill = category_fill
                        ws.cell(row=current_row, column=col).font = category_font
                        ws.cell(row=current_row, column=col).border = border
                    current_row += 1
                    
                    # Gastos de la categoría
                    category_sum = 0.0
                    for expense in expenses:
                        ws.cell(row=current_row, column=1).value = expense.expense_date.strftime('%d/%m/%Y')
                        ws.cell(row=current_row, column=2).value = expense.description
                        ws.cell(row=current_row, column=3).value = expense.expense_type
                        ws.cell(row=current_row, column=4).value = float(expense.amount)
                        ws.cell(row=current_row, column=5).value = ""
                        
                        # Formato de celda
                        for col in range(1, 6):
                            ws.cell(row=current_row, column=col).border = border
                            ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='left')
                        
                        # Formato de moneda para la columna de monto
                        ws.cell(row=current_row, column=4).number_format = '"C$"#,##0.00'
                        
                        category_sum += float(expense.amount)
                        current_row += 1
                    
                    # Subtotal de categoría
                    ws.cell(row=current_row, column=2).value = f"Subtotal {category}"
                    ws.cell(row=current_row, column=2).font = Font(bold=True)
                    ws.cell(row=current_row, column=4).value = category_sum
                    ws.cell(row=current_row, column=4).font = Font(bold=True)
                    ws.cell(row=current_row, column=4).number_format = '"C$"#,##0.00'
                    
                    for col in range(1, 6):
                        ws.cell(row=current_row, column=col).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                        ws.cell(row=current_row, column=col).border = border
                    
                    category_totals[category] = category_sum
                    current_row += 2  # Espacio entre categorías
        
        # Resumen final
        current_row += 1
        ws.cell(row=current_row, column=2).value = "TOTAL GENERAL"
        ws.cell(row=current_row, column=2).font = Font(bold=True, size=12)
        ws.cell(row=current_row, column=4).value = sum(category_totals.values())
        ws.cell(row=current_row, column=4).font = Font(bold=True, size=12)
        ws.cell(row=current_row, column=4).number_format = '"C$"#,##0.00'
        
        for col in range(1, 6):
            ws.cell(row=current_row, column=col).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            ws.cell(row=current_row, column=col).border = border
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        
        # Guardar archivo
        wb.save(output_path)
        return output_path
